from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import calendar

def create_task_workbook():
    wb = Workbook()
    
    # Create the main worksheet
    ws = wb.active
    ws.title = "Tasks"
    
    # Define headers
    headers = ['Date', 'Day', 'Week', 'Task Description', 'Status', 'Priority', 'Category', 'Due Time', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
    
    # Set up dropdown validations
    def create_data_validation(type, formula1):
        return DataValidation(type=type, formula1=formula1, allow_blank=True)

    status_dv = create_data_validation("list", '"Not Started,In Progress,Completed,Delayed,Cancelled"')
    priority_dv = create_data_validation("list", '"High,Medium,Low"')
    category_dv = create_data_validation("list", '"Work,Personal,Health,Family,Finance,Education,Other"')
    
    # Add validations to the main worksheet
    ws.add_data_validation(status_dv)
    ws.add_data_validation(priority_dv)
    ws.add_data_validation(category_dv)
    status_dv.add('E2:E1000')
    priority_dv.add('F2:F1000')
    category_dv.add('G2:G1000')
    
    # Set column widths
    column_widths = [15, 12, 8, 40, 15, 10, 15, 10, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Generate dates for the year
    current_row = 2
    year = 2024
    
    # Recurring tasks templates
    daily_tasks = [
        ("Daily Team Standup", "High", "Work", "09:00"),
        ("Check Emails", "Medium", "Work", "09:30"),
        ("Review Tasks", "Medium", "Work", "17:00")
    ]
    
    weekly_tasks = [
        ("Team Meeting", "High", "Work", "14:00"),
        ("Weekly Report", "High", "Work", "16:00"),
        ("Planning Session", "Medium", "Work", "10:00")
    ]
    
    monthly_tasks = [
        ("Monthly Review", "High", "Work", "15:00"),
        ("Budget Update", "High", "Finance", "11:00"),
        ("Team Assessment", "Medium", "Work", "14:00")
    ]
    
    for month in range(1, 13):
        # Create monthly worksheet
        monthly_ws = wb.create_sheet(title=calendar.month_name[month])
        
        # Copy headers to monthly worksheet
        for col, header in enumerate(headers, 1):
            cell = monthly_ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Create new data validations for the monthly worksheet
        monthly_status_dv = create_data_validation("list", '"Not Started,In Progress,Completed,Delayed,Cancelled"')
        monthly_priority_dv = create_data_validation("list", '"High,Medium,Low"')
        monthly_category_dv = create_data_validation("list", '"Work,Personal,Health,Family,Finance,Education,Other"')
        
        # Add validations to the monthly worksheet
        monthly_ws.add_data_validation(monthly_status_dv)
        monthly_ws.add_data_validation(monthly_priority_dv)
        monthly_ws.add_data_validation(monthly_category_dv)
        monthly_status_dv.add('E2:E1000')
        monthly_priority_dv.add('F2:F1000')
        monthly_category_dv.add('G2:G1000')
        
        # Set column widths for the monthly worksheet
        for i, width in enumerate(column_widths, 1):
            monthly_ws.column_dimensions[chr(64 + i)].width = width
        
        # Fill dates for the month
        num_days = calendar.monthrange(year, month)[1]
        monthly_row = 2
        
        for day in range(1, num_days + 1):
            date = datetime(year, month, day)
            
            # Add daily tasks
            for task, priority, category, time in daily_tasks:
                for ws in [ws, monthly_ws]:
                    ws.cell(row=current_row, column=1, value=date)
                    ws.cell(row=current_row, column=2, value=date.strftime('%A'))
                    ws.cell(row=current_row, column=3, value=date.isocalendar()[1])
                    ws.cell(row=current_row, column=4, value=task)
                    ws.cell(row=current_row, column=5, value="Not Started")
                    ws.cell(row=current_row, column=6, value=priority)
                    ws.cell(row=current_row, column=7, value=category)
                    ws.cell(row=current_row, column=8, value=time)
                current_row += 1
                monthly_row += 1
            
            # Add weekly tasks for Mondays
            if date.weekday() == 0:  # Monday
                for task, priority, category, time in weekly_tasks:
                    for ws in [ws, monthly_ws]:
                        ws.cell(row=current_row, column=1, value=date)
                        ws.cell(row=current_row, column=2, value=date.strftime('%A'))
                        ws.cell(row=current_row, column=3, value=date.isocalendar()[1])
                        ws.cell(row=current_row, column=4, value=task)
                        ws.cell(row=current_row, column=5, value="Not Started")
                        ws.cell(row=current_row, column=6, value=priority)
                        ws.cell(row=current_row, column=7, value=category)
                        ws.cell(row=current_row, column=8, value=time)
                    current_row += 1
                    monthly_row += 1
            
            # Add monthly tasks for first day of the month
            if day == 1:
                for task, priority, category, time in monthly_tasks:
                    for ws in [ws, monthly_ws]:
                        ws.cell(row=current_row, column=1, value=date)
                        ws.cell(row=current_row, column=2, value=date.strftime('%A'))
                        ws.cell(row=current_row, column=3, value=date.isocalendar()[1])
                        ws.cell(row=current_row, column=4, value=task)
                        ws.cell(row=current_row, column=5, value="Not Started")
                        ws.cell(row=current_row, column=6, value=priority)
                        ws.cell(row=current_row, column=7, value=category)
                        ws.cell(row=current_row, column=8, value=time)
                    current_row += 1
                    monthly_row += 1
    
    # Save the workbook
    wb.save('Task_Management_2024.xlsx')

if __name__ == "__main__":
    create_task_workbook()
