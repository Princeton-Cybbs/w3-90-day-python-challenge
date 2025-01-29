import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import (BarChart, Reference, LineChart, PieChart, 
                           ScatterChart, BubbleChart, RadarChart)
from openpyxl.chart.series import DataPoint
import datetime
from datetime import datetime, timedelta
import shutil
import os
import schedule
from plyer import notification

class EnhancedTaskManager:
    def __init__(self, filename="TaskManager2025.xlsx"):
        self.filename = filename
        self.status_options = ["Not Started", "In Progress", "Completed"]
        self.progress_options = ["Pending", "Done"]
        self.priority_colors = {
            "High": "FF0000",    # Red
            "Medium": "FFA500",  # Orange
            "Low": "008000"      # Green
        }
        self.categories = [
            "Core Learning",
            "Personal Development",
            "Essential Activities",
            "Flexible Tasks",
            "Project Work",
            "Health & Wellness",
            "Meetings/Collaboration",
            "Research & Study"
        ]
        
        self.fixed_tasks = [
            ("Project work", "Core Learning", "2:00", "High"),
            ("Cybersecurity practice", "Core Learning", "2:00", "High"),
            ("Reading/revision of slides", "Core Learning", "3:00", "High"),
            ("Huawei practice", "Core Learning", "1:00", "High"),
            ("90 day python", "Core Learning", "0:30", "High"),
            ("Reading self-help books", "Personal Development", "1:00", "Medium"),
            ("Reading the bible", "Personal Development", "0:30", "Medium"),
            ("Duolingo and typing practice", "Personal Development", "0:10", "Medium"),
            ("Sleep", "Essential Activities", "5:00", "High"),
            ("Meals", "Essential Activities", "1:30", "High"),
            ("Personal care", "Essential Activities", "1:00", "High"),
            ("Rest/break periods", "Essential Activities", "2:00", "Medium")
        ]

    def create_monthly_sheet(self, wb, month, year=2025):
        """Create a sheet for a specific month with fixed tasks and time blocks"""
        month_name = datetime(year, month, 1).strftime("%B")
        ws = wb.create_sheet(month_name)
        
        # Define headers
        headers = [
            'Date',
            'Time Block',
            'Duration',
            'Task Description',
            'Category',
            'Status',
            'Progress',
            'Priority',
            'Notes'
        ]
        
        self._setup_sheet_formatting(ws, headers)
        self._add_data_validations(ws)
        self._populate_daily_schedule(ws, month, year)
        self._create_monthly_dashboard(ws)
        self._create_weekly_progress_section(ws)
        
        return ws

    def _setup_sheet_formatting(self, ws, headers):
        """Set up sheet formatting with improved styles"""
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply headers and styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _populate_daily_schedule(self, ws, month, year):
        """Populate daily schedule with fixed tasks and flexible time blocks"""
        current_row = 2
        start_date = datetime(year, month, 1)
        days_in_month = (datetime(year, month % 12 + 1, 1) if month < 12 
                        else datetime(year + 1, 1, 1)) - start_date
        
        for day in range(days_in_month.days):
            current_date = start_date + timedelta(days=day)
            
            # Add fixed tasks
            for task, category, duration, priority in self.fixed_tasks:
                ws.cell(row=current_row, column=1, value=current_date).number_format = 'YYYY-MM-DD'
                ws.cell(row=current_row, column=2, value="Fixed")
                ws.cell(row=current_row, column=3, value=duration)
                ws.cell(row=current_row, column=4, value=task)
                ws.cell(row=current_row, column=5, value=category)
                ws.cell(row=current_row, column=8, value=priority)
                current_row += 1
            
            # Add flexible time blocks
            for block in range(3):  # 3 flexible time blocks of 2 hours each
                ws.cell(row=current_row, column=1, value=current_date).number_format = 'YYYY-MM-DD'
                ws.cell(row=current_row, column=2, value=f"Flexible Block {block + 1}")
                ws.cell(row=current_row, column=3, value="2:00")
                ws.cell(row=current_row, column=5, value="Flexible Tasks")
                current_row += 1

    def _create_monthly_dashboard(self, ws):
        """Create monthly dashboard with enhanced visualizations"""
        dashboard_row = ws.max_row + 5
        
        # Dashboard title
        ws.merge_cells(f'A{dashboard_row}:I{dashboard_row}')
        title_cell = ws.cell(row=dashboard_row, column=1, value="Monthly Progress Dashboard")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Create visualizations
        self._add_task_completion_chart(ws, dashboard_row + 2)
        self._add_time_allocation_chart(ws, dashboard_row + 2)
        self._add_category_progress_chart(ws, dashboard_row + 2)
        self._add_kpi_metrics(ws, dashboard_row + 2)

    def _create_weekly_progress_section(self, ws):
        """Create weekly progress tracking section"""
        weekly_row = ws.max_row + 5
        
        # Weekly progress title
        ws.merge_cells(f'A{weekly_row}:I{weekly_row}')
        title_cell = ws.cell(row=weekly_row, column=1, value="Weekly Progress Tracking")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Add weekly metrics and mini-charts
        self._add_weekly_metrics(ws, weekly_row + 2)
        self._add_weekly_charts(ws, weekly_row + 2)

    def create_yearly_dashboard(self, wb):
        """Create comprehensive yearly dashboard"""
        ws = wb.create_sheet("Yearly Dashboard", 0)
        
        # Dashboard sections
        self._create_yearly_kpi_section(ws)
        self._create_yearly_trends_section(ws)
        self._create_yearly_category_analysis(ws)
        self._create_productivity_patterns(ws)
        
        return ws

    def _create_yearly_kpi_section(self, ws):
        """Create yearly KPI tracking section"""
        kpis = [
            "Overall Task Completion Rate",
            "Core Learning Progress",
            "Personal Development Score",
            "Time Utilization Efficiency",
            "Consistency Score"
        ]
        
        for i, kpi in enumerate(kpis, 2):
            ws.cell(row=i, column=1, value=kpi)
            # Add formulas for KPI calculations
            ws.cell(row=i, column=2, value="=YOURFORMULA")  # Replace with actual formulas

    def create_excel_template(self):
        """Create the complete Excel template"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create yearly dashboard
        self.create_yearly_dashboard(wb)
        
        # Create monthly sheets
        for month in range(1, 13):
            self.create_monthly_sheet(wb, month)
        
        # Save and backup
        wb.save(self.filename)
        self.create_backup()

    def create_backup(self):
        """Create backup with timestamp"""
        backup_dir = "backups"
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = f"{backup_dir}/TaskManager2025_backup_{timestamp}.xlsx"
        shutil.copy2(self.filename, backup_file)

def main():
    task_system = EnhancedTaskManager()
    
    if not os.path.exists(task_system.filename):
        print("Creating new enhanced task management template...")
        task_system.create_excel_template()
        print(f"Template created successfully: {task_system.filename}")
    
    # Schedule automatic backups
    schedule.every().day.at("23:59").do(task_system.create_backup)
    
    print("Setup complete. The system includes:")
    print("- Monthly sheets with fixed and flexible time blocks")
    print("- Monthly and weekly progress tracking")
    print("- Comprehensive yearly dashboard")
    print("- Automatic daily backups")

if __name__ == "__main__":
    main()