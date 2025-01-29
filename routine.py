import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
import datetime
from datetime import datetime, timedelta
import shutil
import os
import schedule
from plyer import notification

class EnhancedTaskManager:
    def __init__(self, filename="TaskManager2025.xlsx"):
        self.filename = filename
        self.status_options = ["Not Started", "In Progress", "Completed"]
        self.progress_options = ["Pending", "Done"]
        self.priority_colors = {
            "High": "FF0000",    # Red
            "Medium": "FFA500",  # Orange
            "Low": "008000"      # Green
        }
        self.task_categories = [
            "Work",
            "Personal",
            "Health",
            "Learning",
            "Finance",
            "Family",
            "Other"
        ]
        
    def create_backup(self):
        """Create backup of the Excel file"""
        backup_dir = "backups"
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = f"{backup_dir}/TaskManager2025_backup_{timestamp}.xlsx"
        shutil.copy2(self.filename, backup_file)
        
    def create_monthly_sheet(self, wb, month, year=2025):
        """Create a sheet for a specific month"""
        month_name = datetime(year, month, 1).strftime("%B")
        ws = wb.create_sheet(month_name)
        
        # Define headers
        headers = [
            'Date',
            'Time',
            'Task Type',
            'Task Description',
            'Category',
            'Status',
            'Progress',
            'Priority',
            'Tags'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply headers and styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Add data validations
        self._add_data_validations(ws)
        
        # Pre-fill dates for the month
        start_date = datetime(year, month, 1)
        next_month = month + 1 if month < 12 else 1
        next_year = year if month < 12 else year + 1
        end_date = datetime(next_year, next_month, 1) - timedelta(days=1)
        
        current_date = start_date
        row = 2
        while current_date <= end_date:
            ws.cell(row=row, column=1, value=current_date).number_format = 'YYYY-MM-DD'
            ws.cell(row=row, column=2, value='09:00').number_format = 'HH:MM'
            row += 1
            current_date += timedelta(days=1)
            
        # Add progress dashboard
        self._create_progress_dashboard(ws, row + 2)
        
        return ws
    
    def _add_data_validations(self, ws):
        """Add data validations to the worksheet"""
        validations = {
            'Status': (self.status_options, 'F'),
            'Progress': (self.progress_options, 'G'),
            'Task Type': (['Daily', 'Weekly', 'Monthly'], 'C'),
            'Priority': (list(self.priority_colors.keys()), 'H'),
            'Category': (self.task_categories, 'E')
        }
        
        for name, (options, column) in validations.items():
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(options)}"',
                allow_blank=True
            )
            ws.add_data_validation(dv)
            dv.add(f'{column}2:{column}1048576')
    
    def _create_progress_dashboard(self, ws, start_row):
        """Create progress dashboard with charts"""
        # Add dashboard title
        dashboard_title = ws.cell(row=start_row, column=1, value="Progress Dashboard")
        dashboard_title.font = Font(size=14, bold=True)
        
        # Create charts
        self._create_task_completion_chart(ws, start_row + 2)
        self._create_priority_distribution_chart(ws, start_row + 2)
        self._create_category_breakdown_chart(ws, start_row + 2)
        
    def _create_task_completion_chart(self, ws, row):
        """Create task completion chart"""
        chart = BarChart()
        chart.title = "Task Completion by Type"
        chart.y_axis.title = "Number of Tasks"
        chart.x_axis.title = "Task Type"
        
        # Add chart to worksheet
        ws.add_chart(chart, f"A{row}")
        
    def _create_priority_distribution_chart(self, ws, row):
        """Create priority distribution chart"""
        chart = PieChart()
        chart.title = "Tasks by Priority"
        
        # Add chart to worksheet
        ws.add_chart(chart, f"F{row}")
        
    def _create_category_breakdown_chart(self, ws, row):
        """Create category breakdown chart"""
        chart = BarChart()
        chart.title = "Tasks by Category"
        chart.y_axis.title = "Number of Tasks"
        chart.x_axis.title = "Category"
        
        # Add chart to worksheet
        ws.add_chart(chart, f"K{row}")
    
    def create_excel_template(self):
        """Create the main Excel template"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create monthly sheets
        for month in range(1, 13):
            self.create_monthly_sheet(wb, month)
        
        # Create summary dashboard
        self._create_summary_dashboard(wb)
        
        # Save the workbook
        wb.save(self.filename)
        
        # Create initial backup
        self.create_backup()
        
    def _create_summary_dashboard(self, wb):
        """Create yearly summary dashboard"""
        ws = wb.create_sheet("Yearly Summary", 0)
        
        # Add summary title
        ws.cell(row=1, column=1, value="2025 Task Management Summary")
        ws.cell(row=1, column=1).font = Font(size=16, bold=True)
        
        # Add summary metrics
        metrics = [
            "Total Tasks",
            "Completed Tasks",
            "Completion Rate",
            "Tasks by Priority",
            "Tasks by Category",
            "Monthly Progress"
        ]
        
        for i, metric in enumerate(metrics, 2):
            ws.cell(row=i, column=1, value=metric)
            
        # Add charts for yearly overview
        self._create_yearly_charts(ws)
    
    def _create_yearly_charts(self, ws):
        """Create charts for yearly summary"""
        # Monthly progress line chart
        chart = LineChart()
        chart.title = "Monthly Task Completion Trend"
        chart.y_axis.title = "Tasks Completed"
        chart.x_axis.title = "Month"
        ws.add_chart(chart, "A15")
        
        # Overall category distribution
        chart = PieChart()
        chart.title = "Yearly Category Distribution"
        ws.add_chart(chart, "H15")

def main():
    task_system = EnhancedTaskManager()
    
    # Create the Excel template if it doesn't exist
    if not os.path.exists(task_system.filename):
        print("Creating new enhanced task management template...")
        task_system.create_excel_template()
        print(f"Template created successfully: {task_system.filename}")
    
    # Schedule automatic backups
    schedule.every().day.at("23:59").do(task_system.create_backup)

if __name__ == "__main__":
    main()