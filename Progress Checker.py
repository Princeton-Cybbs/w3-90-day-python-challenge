import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from datetime import datetime, timedelta
import os

class TaskManagerExcel:
    def __init__(self, filename="TaskManager2025.xlsx"):
        self.filename = filename
        self.status_options = ["Not Started", "In Progress", "Completed"]
        self.categories = [
            "Core Learning",
            "Personal Development",
            "Essential Activities",
            "Flexible Tasks"
        ]
        self.fixed_tasks = [
            ("Project work", "Core Learning", "2:00", "High"),
            ("Cybersecurity practice", "Core Learning", "2:00", "High"),
            ("Reading/revision of slides", "Core Learning", "3:00", "High"),
            ("Huawei practice", "Core Learning", "1:00", "High"),
            ("90 day python", "Core Learning", "0:30", "High"),
            ("Reading self-help books", "Personal Development", "1:00", "Medium"),
            ("Reading the bible", "Personal Development", "0:30", "Medium"),
            ("Duolingo and typing practice", "Personal Development", "0:10", "Medium"),
            ("Sleep", "Essential Activities", "5:00", "High"),
            ("Meals", "Essential Activities", "1:30", "High"),
            ("Personal care", "Essential Activities", "1:00", "High"),
            ("Rest/break periods", "Essential Activities", "2:00", "Medium")
        ]

    def create_workbook(self):
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create yearly dashboard first
        self.create_yearly_dashboard(wb)
        
        # Create monthly sheets
        for month in range(1, 13):
            self.create_monthly_sheet(wb, month)
        
        return wb

    def create_monthly_sheet(self, wb, month):
        month_name = datetime(2025, month, 1).strftime("%B")
        ws = wb.create_sheet(month_name)
        
        # Setup headers
        headers = [
            'Date',
            'Time Block',
            'Duration',
            'Task Description',
            'Category',
            'Status',
            'Progress',
            'Priority',
            'Notes'
        ]
        
        # Style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Add data validation
        status_dv = DataValidation(type="list", formula1=f'"{",".join(self.status_options)}"')
        category_dv = DataValidation(type="list", formula1=f'"{",".join(self.categories)}"')
        progress_dv = DataValidation(type="list", formula1='"Pending,Done"')
        priority_dv = DataValidation(type="list", formula1='"High,Medium,Low"')

        ws.add_data_validation(status_dv)
        ws.add_data_validation(category_dv)
        ws.add_data_validation(progress_dv)
        ws.add_data_validation(priority_dv)

        # Apply validations
        status_dv.add(f'F2:F1000')
        category_dv.add(f'E2:E1000')
        progress_dv.add(f'G2:G1000')
        priority_dv.add(f'H2:H1000')

        # Populate daily tasks
        self.populate_monthly_tasks(ws, month)
        
        # Add monthly dashboard
        self.add_monthly_dashboard(ws)
        
        return ws

    def populate_monthly_tasks(self, ws, month):
        current_row = 2
        start_date = datetime(2025, month, 1)
        end_date = (datetime(2025, month + 1, 1) if month < 12 
                   else datetime(2026, 1, 1))
        
        current_date = start_date
        while current_date < end_date:
            # Add fixed tasks
            for task, category, duration, priority in self.fixed_tasks:
                ws.cell(row=current_row, column=1, value=current_date)
                ws.cell(row=current_row, column=2, value="Fixed")
                ws.cell(row=current_row, column=3, value=duration)
                ws.cell(row=current_row, column=4, value=task)
                ws.cell(row=current_row, column=5, value=category)
                ws.cell(row=current_row, column=8, value=priority)
                current_row += 1
            
            # Add flexible blocks
            for block in range(3):
                ws.cell(row=current_row, column=1, value=current_date)
                ws.cell(row=current_row, column=2, value=f"Flexible Block {block + 1}")
                ws.cell(row=current_row, column=3, value="2:00")
                ws.cell(row=current_row, column=5, value="Flexible Tasks")
                current_row += 1
            
            current_date += timedelta(days=1)

    def add_monthly_dashboard(self, ws):
        # Add dashboard section after tasks
        dashboard_row = ws.max_row + 5
        
        ws.cell(row=dashboard_row, column=1, value="Monthly Dashboard")
        ws.cell(row=dashboard_row, column=1).font = Font(bold=True, size=14)
        
        # Add completion rate chart
        completion_chart = PieChart()
        completion_chart.title = "Task Completion Rate"
        ws.add_chart(completion_chart, f"A{dashboard_row + 2}")
        
        # Add category distribution chart
        category_chart = BarChart()
        category_chart.title = "Tasks by Category"
        ws.add_chart(category_chart, f"H{dashboard_row + 2}")

    def create_yearly_dashboard(self, wb):
        ws = wb.create_sheet("Yearly Dashboard", 0)
        
        # Add title
        ws.cell(row=1, column=1, value="2025 Task Management Dashboard")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        
        # Add KPI sections
        kpi_titles = [
            "Overall Completion Rate",
            "Core Learning Progress",
            "Personal Development Score",
            "Time Utilization",
            "Consistency Score"
        ]
        
        for i, title in enumerate(kpi_titles, 2):
            ws.cell(row=i, column=1, value=title)
            ws.cell(row=i, column=2, value="0%")  # Default value
        
        # Add yearly trend chart
        trend_chart = LineChart()
        trend_chart.title = "Monthly Progress Trend"
        ws.add_chart(trend_chart, "A10")

    def generate_excel(self):
        wb = self.create_workbook()
        wb.save(self.filename)
        print(f"Task management Excel file created: {self.filename}")

if __name__ == "__main__":
    task_manager = TaskManagerExcel()
    task_manager.generate_excel()